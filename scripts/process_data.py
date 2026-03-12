"""
Data Processing Script — Reshoring Intelligence App
Downloads and merges:
  1. World Bank Governance Indicators (WGI) — political stability, rule of law, etc.
  2. World Bank Economic Data — GDP, growth, population, FDI
  3. Transparency International CPI 2024 — corruption perceptions
  4. B-READY 2025 pillar scores (local Excel files)
Output: src/data/countries.json
"""

import requests
import json
import time
import os
import warnings
warnings.filterwarnings("ignore")

import openpyxl

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE        = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR    = os.path.join(BASE, "src", "data")
BREADY      = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "B-READY-2025-PILLAR-TOPIC-SCORES.xlsx")
WTO_API_KEY = os.environ.get("WTO_API_KEY")   # optional — see scripts/WTO_SETUP.md

os.makedirs(DATA_DIR, exist_ok=True)

# ── Helpers ───────────────────────────────────────────────────────────────────

def wb_fetch(indicator, date="2020:2024"):
    """Fetch a World Bank indicator for all countries. Returns (dict, median_year)."""
    url = (f"https://api.worldbank.org/v2/country/all/indicator/{indicator}"
           f"?format=json&date={date}&per_page=300&mrv=1")
    try:
        r = requests.get(url, timeout=15)
        r.raise_for_status()
        data = r.json()
        if len(data) < 2:
            return {}, None
        result = {}
        years = []
        for item in data[1]:
            if item.get("countryiso3code") and item.get("value") is not None:
                result[item["countryiso3code"]] = item["value"]
                try:
                    years.append(int(item["date"]))
                except:
                    pass
        median_year = sorted(years)[len(years)//2] if years else None
        return result, median_year
    except Exception as e:
        print(f"  Warning: could not fetch {indicator}: {e}")
        return {}, None

def normalize_wgi(val):
    """Normalize WGI score from -2.5..2.5 to 0..100."""
    if val is None:
        return None
    return round((val + 2.5) / 5.0 * 100, 1)

# ── Step 1: B-READY Pillar Scores ─────────────────────────────────────────────
print("Loading B-READY 2025 data...")

wb_bready = openpyxl.load_workbook(BREADY, data_only=True)

# Master pillar scores (3 pillars)
ws_master = wb_bready["00_B-READY_Pillar_Score"]
bready = {}
rows = list(ws_master.iter_rows(values_only=True))
for row in rows[1:]:
    if row[0] and row[1]:
        economy, code = row[0], row[1]
        bready[code] = {
            "name": economy,
            "code": code,
            "bready_regulatory_framework": round(row[2], 1) if row[2] else None,
            "bready_public_services":      round(row[3], 1) if row[3] else None,
            "bready_operational_efficiency": round(row[4], 1) if row[4] else None,
        }

# Individual pillar sheets — extract economy-level topic scores
pillar_sheets = {
    "01_Business_Entry":       "bready_business_entry",
    "02_Business_Location":    "bready_business_location",
    "03_Utility_Services":     "bready_utility_services",
    "04_Labor":                "bready_labor",
    "05_Financial_Services":   "bready_financial_services",
    "06_International_Trade":  "bready_international_trade",
    "07_Taxation":             "bready_taxation",
    "08_Dispute_Resolution":   "bready_dispute_resolution",
    "09_Market_Competition":   "bready_market_competition",
    "10_Business_Insolvency":  "bready_business_insolvency",
}

for sheet_name, key in pillar_sheets.items():
    try:
        ws = wb_bready[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # Find header row with economy names (look for 'Economy' in first column)
        header_row_idx = None
        for i, row in enumerate(rows):
            if row[0] == "Economy":
                header_row_idx = i
                break
        if header_row_idx is None:
            continue
        header = rows[header_row_idx]
        # Find the "Pillar Score" or composite score column
        # The first numeric column after 'Economy Code' is the overall pillar score
        for data_row in rows[header_row_idx + 1:]:
            if not data_row[0] or not data_row[1]:
                continue
            code = data_row[1]
            # Find first numeric score — pillar composite
            score = None
            for val in data_row[2:]:
                if isinstance(val, (int, float)):
                    score = round(float(val), 1)
                    break
            if code in bready and score is not None:
                bready[code][key] = score
    except Exception as e:
        print(f"  Warning: {sheet_name}: {e}")

print(f"  Loaded {len(bready)} countries from B-READY")

# ── Step 2: World Bank Governance Indicators ──────────────────────────────────
print("Fetching World Bank Governance Indicators...")

wgi_indicators = {
    "PV.EST": "wgi_political_stability",
    "GE.EST": "wgi_govt_effectiveness",
    "RQ.EST": "wgi_regulatory_quality",
    "RL.EST": "wgi_rule_of_law",
    "CC.EST": "wgi_corruption_control",
    "VA.EST": "wgi_voice_accountability",
}

wgi_data = {}
wgi_years = {}
for indicator, key in wgi_indicators.items():
    print(f"  Fetching {indicator}...")
    raw, yr = wb_fetch(indicator, date="2019:2023")
    wgi_data[key] = raw
    if yr: wgi_years[key] = yr
    time.sleep(0.3)

# Collect all country codes across WGI + B-READY
all_codes = set(bready.keys())
for key, raw in wgi_data.items():
    all_codes.update(raw.keys())

# ── Step 3: World Bank Economic Indicators ────────────────────────────────────
print("Fetching World Bank economic data...")

econ_indicators = {
    "NY.GDP.MKTP.CD":        "gdp_usd",
    "NY.GDP.MKTP.KD.ZG":     "gdp_growth_pct",
    "NY.GDP.PCAP.CD":        "gdp_per_capita",
    "SP.POP.TOTL":           "population",
    "BX.KLT.DINV.WD.GD.ZS": "fdi_pct_gdp",
    "IC.BUS.EASE.XQ":        "ease_of_business_rank",
}

econ_data = {}
econ_years = {}
for indicator, key in econ_indicators.items():
    print(f"  Fetching {indicator}...")
    raw, yr = wb_fetch(indicator, date="2020:2024")
    econ_data[key] = raw
    if yr: econ_years[key] = yr
    time.sleep(0.3)

# ── Step 3.5: Trade & Logistics Indicators ────────────────────────────────────
print("Fetching trade and logistics data...")

trade_indicators = {
    "TM.TAX.MRCH.WM.AR.ZS": "tariff_rate_weighted_mean",   # applied tariff, weighted mean %
    "TM.TAX.MRCH.SM.AR.ZS": "tariff_rate_simple_mean",     # applied tariff, simple mean %
    "NE.TRD.GNFS.ZS":       "trade_pct_gdp",               # trade as % of GDP
    "TX.VAL.MRCH.WD.CD":    "merchandise_exports_usd",     # merchandise exports USD
    "TM.VAL.MRCH.WD.CD":    "merchandise_imports_usd",     # merchandise imports USD
    "LP.LPI.OVRL.XQ":       "logistics_performance_index", # WB LPI (1-5 scale)
}

trade_data = {}
trade_years = {}
for indicator, key in trade_indicators.items():
    print(f"  Fetching {indicator}...")
    raw, yr = wb_fetch(indicator, date="2018:2024")
    trade_data[key] = raw
    if yr: trade_years[key] = yr
    time.sleep(0.3)

# ── Step 3.6: WTO Bound Tariff Data (optional) ────────────────────────────────
wto_bound_data = {}

if WTO_API_KEY:
    print("Fetching WTO bound tariff data (API key found)...")

    def wto_fetch_bound(codes, year=2022):
        """Fetch WTO simple-mean bound tariff for a batch of countries."""
        batch = ",".join(codes)
        url = (
            f"https://api.wto.org/statistics/tariff/bound"
            f"?reporterCode={batch}&year={year}&format=json"
        )
        try:
            r = requests.get(url, timeout=20,
                             headers={"Authorization": f"Bearer {WTO_API_KEY}"})
            if r.status_code == 200:
                result = {}
                for item in r.json().get("Dataset", []):
                    c = item.get("reporterISO3")
                    v = item.get("value")
                    if c and v is not None:
                        result[str(c)] = float(v)
                return result
        except Exception as e:
            print(f"  Warning: WTO fetch failed: {e}")
        return {}

    # Batch countries to stay within rate limits (600 req/day free tier)
    all_codes_list = list(all_codes)
    for i in range(0, len(all_codes_list), 15):
        batch = all_codes_list[i:i+15]
        wto_bound_data.update(wto_fetch_bound(batch))
        time.sleep(0.5)

    print(f"  {len(wto_bound_data)} countries have WTO bound tariff data")
else:
    print("  WTO_API_KEY not set — skipping WTO bound tariffs (see scripts/WTO_SETUP.md)")

# Also fetch country metadata (region, income level)
print("  Fetching country metadata...")
try:
    meta_url = "https://api.worldbank.org/v2/country?format=json&per_page=300"
    r = requests.get(meta_url, timeout=15)
    meta_raw = r.json()[1]
    country_meta = {
        c["id"]: {
            "region": c.get("region", {}).get("value", ""),
            "income_level": c.get("incomeLevel", {}).get("value", ""),
            "capital": c.get("capitalCity", ""),
            "longitude": c.get("longitude", ""),
            "latitude":  c.get("latitude", ""),
        }
        for c in meta_raw
        if c.get("id") and c.get("region", {}).get("id") != "NA"
    }
except Exception as e:
    print(f"  Warning: metadata fetch failed: {e}")
    country_meta = {}

# ── Step 4: Transparency International CPI 2024 ───────────────────────────────
print("Downloading Transparency International CPI 2024...")

cpi_data = {}
cpi_urls = [
    "https://images.transparencycdn.org/images/CPI2024_Country-Results-and-Sources.xlsx",
    "https://www.transparency.org/files/content/pages/CPI2024_Country-Results-and-Sources.xlsx",
]

cpi_downloaded = False
for url in cpi_urls:
    try:
        r = requests.get(url, timeout=20)
        if r.status_code == 200 and len(r.content) > 10000:
            cpi_path = os.path.join(DATA_DIR, "CPI2024.xlsx")
            with open(cpi_path, "wb") as f:
                f.write(r.content)
            wb_cpi = openpyxl.load_workbook(cpi_path, data_only=True)
            ws_cpi = wb_cpi.active
            cpi_rows = list(ws_cpi.iter_rows(values_only=True))
            # Find header row
            for i, row in enumerate(cpi_rows):
                if row and any(str(v).upper() in ("ISO3", "ISO", "COUNTRY CODE") for v in row if v):
                    header = [str(v).strip() if v else "" for v in row]
                    iso_col  = next((j for j, h in enumerate(header) if "ISO" in h.upper()), None)
                    score_col = next((j for j, h in enumerate(header) if "CPI SCORE" in h.upper() or "SCORE 2024" in h.upper()), None)
                    if iso_col is not None and score_col is not None:
                        for data_row in cpi_rows[i+1:]:
                            if data_row[iso_col] and data_row[score_col]:
                                try:
                                    cpi_data[str(data_row[iso_col]).strip()] = float(data_row[score_col])
                                except:
                                    pass
                    break
            if cpi_data:
                cpi_downloaded = True
                print(f"  Loaded {len(cpi_data)} CPI scores")
                break
    except Exception as e:
        print(f"  CPI URL failed: {e}")

if not cpi_downloaded:
    print("  CPI download failed — will use WGI corruption control as proxy")

# ── Step 5: Corporate Tax Rates (static table for key countries) ───────────────
print("Loading corporate tax rates...")

# OECD + key non-OECD corporate tax rates (2024)
corporate_tax = {
    "ARG": 35.0, "AUS": 30.0, "AUT": 23.0, "BEL": 25.0, "BRA": 34.0,
    "CAN": 26.5, "CHL": 27.0, "CHN": 25.0, "COL": 35.0, "CZE": 21.0,
    "DEU": 29.9, "DNK": 22.0, "EGY": 22.5, "ESP": 25.0, "EST": 20.0,
    "FIN": 20.0, "FRA": 25.0, "GBR": 25.0, "GRC": 22.0, "HUN": 9.0,
    "IDN": 22.0, "IND": 25.2, "IRL": 12.5, "ISR": 23.0, "ITA": 24.0,
    "JPN": 29.7, "KOR": 24.0, "LTU": 15.0, "LUX": 17.0, "LVA": 20.0,
    "MEX": 30.0, "MYS": 24.0, "NLD": 25.8, "NOR": 22.0, "NZL": 28.0,
    "PHL": 25.0, "POL": 19.0, "PRT": 21.0, "ROU": 16.0, "SAU": 20.0,
    "SGP": 17.0, "SVK": 21.0, "SVN": 19.0, "SWE": 20.6, "THA": 20.0,
    "TUR": 25.0, "TWN": 20.0, "USA": 21.0, "VNM": 20.0, "ZAF": 27.0,
    "ARE": 9.0,  "ARM": 18.0, "AZE": 20.0, "BGD": 27.5, "BGR": 10.0,
    "BHR": 0.0,  "BWA": 22.0, "CMR": 33.0, "CRI": 30.0, "DOM": 27.0,
    "DZA": 26.0, "ECU": 25.0, "ETH": 30.0, "GEO": 15.0, "GHA": 25.0,
    "GTM": 25.0, "HND": 30.0, "HRV": 18.0, "JOR": 20.0, "KAZ": 20.0,
    "KEN": 30.0, "KGZ": 10.0, "KHM": 20.0, "LKA": 30.0, "MAR": 31.0,
    "MDG": 20.0, "MLI": 30.0, "MMR": 25.0, "MOZ": 32.0, "MUS": 15.0,
    "MWI": 30.0, "NAM": 32.0, "NGA": 30.0, "NPL": 25.0, "OMN": 15.0,
    "PAK": 29.0, "PAN": 25.0, "PER": 29.5, "PRY": 10.0, "QAT": 10.0,
    "RWA": 30.0, "SEN": 30.0, "SLV": 30.0, "SRB": 15.0, "TJK": 18.0,
    "TTO": 30.0, "TUN": 15.0, "TZA": 30.0, "UGA": 30.0, "UKR": 18.0,
    "URY": 25.0, "UZB": 15.0, "VEN": 34.0, "XKX": 10.0, "ZMB": 30.0,
    "ZWE": 24.0, "AGO": 25.0, "CIV": 25.0, "COD": 35.0, "HKG": 16.5,
    "BLR": 20.0, "MKD": 10.0, "MDA": 12.0, "ALB": 15.0, "MNE": 9.0,
    "BIH": 10.0, "KWT": 15.0,
}

print(f"  {len(corporate_tax)} corporate tax rates loaded")

# ── Step 6: Merge All Data ────────────────────────────────────────────────────
print("Merging all datasets...")

# Collect all known country codes
all_iso3 = set(bready.keys()) | set(country_meta.keys())

countries = []
for code in all_iso3:
    if not code or len(code) != 3:
        continue

    meta = country_meta.get(code, {})
    b = bready.get(code, {})

    # Skip aggregates (World Bank includes regional aggregates with 3-char codes)
    region = meta.get("region", "")
    if not region and code not in bready:
        continue

    # B-READY composite (average of 3 pillars)
    rf = b.get("bready_regulatory_framework")
    ps = b.get("bready_public_services")
    oe = b.get("bready_operational_efficiency")
    bready_vals = [v for v in [rf, ps, oe] if v is not None]
    bready_composite = round(sum(bready_vals) / len(bready_vals), 1) if bready_vals else None

    # WGI scores (normalized 0-100)
    pol_stab   = normalize_wgi(wgi_data.get("wgi_political_stability", {}).get(code))
    govt_eff   = normalize_wgi(wgi_data.get("wgi_govt_effectiveness", {}).get(code))
    reg_qual   = normalize_wgi(wgi_data.get("wgi_regulatory_quality", {}).get(code))
    rule_law   = normalize_wgi(wgi_data.get("wgi_rule_of_law", {}).get(code))
    corr_ctrl  = normalize_wgi(wgi_data.get("wgi_corruption_control", {}).get(code))
    voice_acc  = normalize_wgi(wgi_data.get("wgi_voice_accountability", {}).get(code))

    # CPI score (0-100, higher = less corrupt)
    cpi_score = cpi_data.get(code)

    # Corruption: prefer TI CPI, fall back to WGI CC
    corruption_score = cpi_score if cpi_score is not None else corr_ctrl

    # Economic data
    gdp         = econ_data.get("gdp_usd", {}).get(code)
    gdp_growth  = econ_data.get("gdp_growth_pct", {}).get(code)
    gdp_pc      = econ_data.get("gdp_per_capita", {}).get(code)
    population  = econ_data.get("population", {}).get(code)
    fdi         = econ_data.get("fdi_pct_gdp", {}).get(code)
    tax_rate    = corporate_tax.get(code)

    tariff_wmean  = trade_data.get("tariff_rate_weighted_mean", {}).get(code)
    tariff_smean  = trade_data.get("tariff_rate_simple_mean", {}).get(code)
    trade_gdp     = trade_data.get("trade_pct_gdp", {}).get(code)
    exports_usd   = trade_data.get("merchandise_exports_usd", {}).get(code)
    imports_usd   = trade_data.get("merchandise_imports_usd", {}).get(code)
    lpi           = trade_data.get("logistics_performance_index", {}).get(code)
    tariff_bound  = wto_bound_data.get(code)   # WTO MFN bound tariff simple mean

    # ── Composite Viability Score (0-100) ──────────────────────────────────
    # Default weights (adjustable in app):
    # Business Environment: 35%, Political Stability: 25%,
    # Rule of Law: 15%, Corruption: 15%, Govt Effectiveness: 10%
    score_components = {
        "business_environment": (bready_composite, 0.35),
        "political_stability":  (pol_stab,         0.25),
        "rule_of_law":          (rule_law,          0.15),
        "corruption_control":   (corruption_score,  0.15),
        "govt_effectiveness":   (govt_eff,          0.10),
    }
    weighted_sum = 0
    weight_total = 0
    for val, wt in score_components.values():
        if val is not None:
            weighted_sum += val * wt
            weight_total += wt
    overall = round(weighted_sum / weight_total, 1) if weight_total > 0.2 else None

    entry = {
        "code":         code,
        "name":         b.get("name") or meta.get("name", code),
        "region":       region,
        "income_level": meta.get("income_level", ""),
        "capital":      meta.get("capital", ""),
        "lat":          meta.get("latitude", ""),
        "lon":          meta.get("longitude", ""),
        # Composite score
        "overall_score": overall,
        # Dimension scores
        "scores": {
            "business_environment": bready_composite,
            "political_stability":  pol_stab,
            "govt_effectiveness":   govt_eff,
            "regulatory_quality":   reg_qual,
            "rule_of_law":          rule_law,
            "corruption_control":   corruption_score,
            "voice_accountability": voice_acc,
        },
        # B-READY pillars
        "bready": {
            "composite":                bready_composite,
            "regulatory_framework":     rf,
            "public_services":          ps,
            "operational_efficiency":   oe,
            "business_entry":           b.get("bready_business_entry"),
            "business_location":        b.get("bready_business_location"),
            "utility_services":         b.get("bready_utility_services"),
            "labor":                    b.get("bready_labor"),
            "financial_services":       b.get("bready_financial_services"),
            "international_trade":      b.get("bready_international_trade"),
            "taxation":                 b.get("bready_taxation"),
            "dispute_resolution":       b.get("bready_dispute_resolution"),
            "market_competition":       b.get("bready_market_competition"),
            "business_insolvency":      b.get("bready_business_insolvency"),
        },
        # Economic data
        "economic": {
            "gdp_usd":          gdp,
            "gdp_growth_pct":   round(gdp_growth, 2) if gdp_growth else None,
            "gdp_per_capita":   round(gdp_pc) if gdp_pc else None,
            "population":       round(population) if population else None,
            "fdi_pct_gdp":      round(fdi, 2) if fdi else None,
            "corporate_tax_rate": tax_rate,
            "cpi_score":        cpi_score,
        },
        "trade": {
            "tariff_rate_weighted_mean": round(tariff_wmean, 1) if tariff_wmean else None,
            "tariff_rate_simple_mean":   round(tariff_smean, 1) if tariff_smean else None,
            "tariff_bound_mean":         round(tariff_bound, 1) if tariff_bound else None,
            "tariff_overhang":           round(tariff_bound - tariff_smean, 1)
                                         if (tariff_bound and tariff_smean) else None,
            "trade_pct_gdp":             round(trade_gdp, 1) if trade_gdp else None,
            "merchandise_exports_usd":   round(exports_usd) if exports_usd else None,
            "merchandise_imports_usd":   round(imports_usd) if imports_usd else None,
            "trade_balance_usd":         round(exports_usd - imports_usd) if (exports_usd and imports_usd) else None,
            "logistics_performance_index": round(lpi, 2) if lpi else None,
        },
    }
    countries.append(entry)

# Sort by overall score descending
countries.sort(key=lambda x: x.get("overall_score") or 0, reverse=True)

out_path = os.path.join(DATA_DIR, "countries.json")
with open(out_path, "w") as f:
    json.dump(countries, f, indent=2)

print(f"\n✓ Done. {len(countries)} countries written to {out_path}")
scored = sum(1 for c in countries if c.get("overall_score") is not None)
print(f"  {scored} countries have a full composite score")
print(f"  {len([c for c in countries if c['bready']['composite']])} have B-READY data")
print(f"  {len([c for c in countries if c['scores']['political_stability']])} have WGI political stability data")
print(f"  {len([c for c in countries if c['economic']['cpi_score']])} have TI CPI scores")
print(f"  {len([c for c in countries if c['economic']['corporate_tax_rate']])} have corporate tax rates")

# ── Step 7: Write Data Freshness Metadata ─────────────────────────────────────
import datetime
freshness = {
    "generated_utc": datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
    "sources": {
        "bready": {
            "label": "World Bank B-READY",
            "year": 2025,
            "coverage": len([c for c in countries if c["bready"]["composite"]]),
            "url": "https://www.worldbank.org/en/programs/business-enabling-environment"
        },
        "wgi": {
            "label": "World Bank Governance Indicators",
            "year": int(sum(wgi_years.values()) / len(wgi_years)) if wgi_years else 2022,
            "coverage": len([c for c in countries if c["scores"]["political_stability"]]),
            "url": "https://info.worldbank.org/governance/wgi/"
        },
        "trade": {
            "label": "World Bank Trade & Tariff Data",
            "year": int(sum(trade_years.values()) / len(trade_years)) if trade_years else 2022,
            "coverage": len([c for c in countries if c["trade"]["tariff_rate_weighted_mean"]]),
            "url": "https://wits.worldbank.org/"
        },
        "lpi": {
            "label": "World Bank Logistics Performance Index",
            "year": trade_years.get("logistics_performance_index", 2023),
            "coverage": len([c for c in countries if c["trade"]["logistics_performance_index"]]),
            "url": "https://lpi.worldbank.org/"
        },
        "economic": {
            "label": "World Bank Economic Indicators",
            "year": int(sum(econ_years.values()) / len(econ_years)) if econ_years else 2023,
            "coverage": len([c for c in countries if c["economic"]["gdp_usd"]]),
            "url": "https://data.worldbank.org/"
        },
    }
}

if wto_bound_data:
    freshness["sources"]["wto"] = {
        "label": "WTO Bound Tariff Database",
        "year": 2022,
        "coverage": len(wto_bound_data),
        "url": "https://apiportal.wto.org/"
    }
freshness_path = os.path.join(DATA_DIR, "data_freshness.json")
with open(freshness_path, "w") as f:
    json.dump(freshness, f, indent=2)
print(f"✓ Data freshness metadata written to {freshness_path}")
